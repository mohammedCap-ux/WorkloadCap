"""
seed_geo.py
-----------
Script de seed pour les donnees stables de geolocalisation.

Lit data/Geolocalisation File.xlsx et remplit les 3 tables :
- geo_suppliers (depuis feuille "Suppliers Location")
- geo_docks (depuis feuille "CSCs Location")
- geo_availability (depuis feuille "Availability")

Re-runnable : vide les 3 tables avant insert. A lancer manuellement :
    .\venv\Scripts\python.exe seed_geo.py

Politique de cleanup :
- Suppliers sans coords valides : SKIP + log warning
- Docks sans coords valides : SKIP + log warning
- Availability : tout importer (les liaisons sont utiles meme sans coords)
"""

import os
import sys
from pathlib import Path

import openpyxl

from database import SessionLocal, engine, Base
from models import GeoSupplier, GeoDock, GeoAvailability


# ============================================================
#   CONFIG
# ============================================================

BASE_DIR = Path(__file__).resolve().parent.parent
XLSX_PATH = BASE_DIR / "data" / "Geolocalisation File.xlsx"


# ============================================================
#   HELPERS
# ============================================================

def to_float(v):
    """Convertit en float, retourne None si invalide."""
    if v is None or v == "":
        return None
    try:
        if isinstance(v, str):
            v = v.replace(",", ".").strip()
        f = float(v)
        if f != f:  # NaN check
            return None
        return f
    except (ValueError, TypeError):
        return None


def clean_str(v):
    """Strip + None si vide."""
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def has_valid_coords(lat, lng):
    """Vrai si lat/lng sont des floats dans une plage plausible."""
    if lat is None or lng is None:
        return False
    if not (-90 <= lat <= 90):
        return False
    if not (-180 <= lng <= 180):
        return False
    if lat == 0 and lng == 0:
        return False  # (0, 0) = coord par defaut suspecte
    return True


def find_sheet(wb, *keywords):
    """Cherche une feuille par mots-cles (case-insensitive)."""
    for kw in keywords:
        kw_lower = kw.lower()
        for name in wb.sheetnames:
            if kw_lower in name.lower():
                return wb[name]
    return None


def headers_map(ws):
    """Retourne {header_norm: col_idx} pour la 1ere ligne (1-indexed)."""
    result = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        # Normalize : lowercase + remove newlines/extra spaces
        norm = str(cell.value).lower().replace("\n", " ").strip()
        norm = " ".join(norm.split())
        result[norm] = col_idx
    return result


def get_cell(row, headers, *candidates):
    """Recupere la valeur d'une cellule en cherchant plusieurs noms de colonne."""
    for c in candidates:
        c_norm = c.lower().replace("\n", " ").strip()
        c_norm = " ".join(c_norm.split())
        if c_norm in headers:
            col_idx = headers[c_norm]
            return row[col_idx - 1].value
    return None


# ============================================================
#   PARSERS
# ============================================================

def parse_suppliers(ws, db):
    """Parse la feuille Suppliers Location."""
    headers = headers_map(ws)
    print(f"\n[suppliers] Headers detectes : {list(headers.keys())[:12]}")

    inserted = 0
    skipped_no_coords = 0
    skipped_total = 0

    for row in ws.iter_rows(min_row=2, values_only=False):
        # Lignes vides ?
        if all(cell.value is None for cell in row):
            continue

        lat = to_float(get_cell(row, headers, "Latitude", "lat"))
        lng = to_float(get_cell(row, headers, "Longitude", "lng", "lon"))

        if not has_valid_coords(lat, lng):
            skipped_no_coords += 1
            skipped_total += 1
            continue

        item = GeoSupplier(
            xf_code=clean_str(get_cell(row, headers, "XF CODE", "XFCode")),
            seller_cofor=clean_str(get_cell(row, headers, "SELLER COFOR")),
            shipper_cofor=clean_str(get_cell(row, headers, "SHIPPER COFOR")),
            empty_return_cofor=clean_str(get_cell(row, headers, "EMPTY RETURN COFOR")),
            supplier_name=clean_str(get_cell(row, headers, "SUPPLIER")),
            address=clean_str(get_cell(row, headers, "Address", "Adresse")),
            city=clean_str(get_cell(row, headers, "City")),
            country=clean_str(get_cell(row, headers, "Country")),
            lat=lat,
            lng=lng,
        )
        db.add(item)
        inserted += 1

    db.commit()
    print(f"[suppliers] Inseres : {inserted}")
    print(f"[suppliers] SKIP (coords invalides) : {skipped_no_coords}")
    return inserted, skipped_total


def parse_docks(ws, db):
    """Parse la feuille CSCs Location."""
    headers = headers_map(ws)
    print(f"\n[docks] Headers detectes : {list(headers.keys())}")

    inserted = 0
    skipped_no_coords = 0
    duplicates_skipped = 0
    seen_names = set()

    for row in ws.iter_rows(min_row=2, values_only=False):
        if all(cell.value is None for cell in row):
            continue

        name = clean_str(get_cell(row, headers, "Dock", "Docks"))
        if not name:
            continue

        if name in seen_names:
            duplicates_skipped += 1
            print(f"[docks] WARNING: doublon de nom -> skip : {name}")
            continue

        lat = to_float(get_cell(row, headers, "Latitude"))
        lng = to_float(get_cell(row, headers, "Longitude"))

        if not has_valid_coords(lat, lng):
            skipped_no_coords += 1
            print(f"[docks] WARNING: coords invalides -> skip : {name}")
            continue

        seen_names.add(name)
        item = GeoDock(
            name=name,
            city=clean_str(get_cell(row, headers, "City")),
            country=clean_str(get_cell(row, headers, "Country")),
            lat=lat,
            lng=lng,
        )
        db.add(item)
        inserted += 1

    db.commit()
    print(f"[docks] Inseres : {inserted}")
    print(f"[docks] SKIP (coords invalides) : {skipped_no_coords}")
    print(f"[docks] SKIP (doublons) : {duplicates_skipped}")
    return inserted, skipped_no_coords + duplicates_skipped


def parse_availability(ws, db):
    """Parse la feuille Availability."""
    headers = headers_map(ws)
    print(f"\n[availability] Headers detectes : {list(headers.keys())[:18]}")

    inserted = 0

    for row in ws.iter_rows(min_row=2, values_only=False):
        if all(cell.value is None for cell in row):
            continue

        # Construire la liste des alternates (1 a 8)
        alternates = []
        for n in range(1, 10):
            v = clean_str(get_cell(row, headers, f"Alternate {n}", f"Alternative {n}"))
            if v:
                alternates.append(v)

        item = GeoAvailability(
            mastercode=clean_str(get_cell(row, headers, "Mastercode")),
            supplier_name=clean_str(get_cell(row, headers, "Name", "Supplier")),
            seller_cofor=clean_str(get_cell(row, headers, "Seller Cofor", "SELLER COFOR")),
            shipper_cofor=clean_str(get_cell(row, headers, "Shipper Cofor", "SHIPPER COFOR")),
            empty_return_cofor=clean_str(get_cell(row, headers, "Empty return", "Empty Return", "EMPTY RETURN COFOR")),
            packaging_id=clean_str(get_cell(row, headers, "Packaging", "Packaging ID")),
            pooling_dock=clean_str(get_cell(row, headers, "CSC", "Pooling dock", "Pooling Dock")),
            alternates=alternates,
        )
        db.add(item)
        inserted += 1

    db.commit()
    print(f"[availability] Inseres : {inserted}")
    return inserted, 0


# ============================================================
#   MAIN
# ============================================================

def main():
    if not XLSX_PATH.exists():
        print(f"[FATAL] Fichier introuvable : {XLSX_PATH}")
        sys.exit(1)

    print(f"Source : {XLSX_PATH}")
    print(f"Taille : {XLSX_PATH.stat().st_size} octets\n")

    # S'assurer que les tables existent
    Base.metadata.create_all(bind=engine)

    # Ouvrir le xlsx
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    print(f"Feuilles disponibles : {wb.sheetnames}\n")

    # Localiser les 3 feuilles utiles
    sheet_suppliers = find_sheet(wb, "supplier", "fournis")
    sheet_docks = find_sheet(wb, "csc", "dock")
    sheet_availability = find_sheet(wb, "availab")

    if not sheet_suppliers:
        print("[FATAL] Feuille 'Suppliers Location' introuvable")
        sys.exit(1)
    if not sheet_docks:
        print("[FATAL] Feuille 'CSCs Location' introuvable")
        sys.exit(1)
    if not sheet_availability:
        print("[FATAL] Feuille 'Availability' introuvable")
        sys.exit(1)

    print(f"Feuille suppliers   : {sheet_suppliers.title}")
    print(f"Feuille docks       : {sheet_docks.title}")
    print(f"Feuille availability: {sheet_availability.title}")

    db = SessionLocal()
    try:
        # Vider les tables (re-runnable)
        print("\n[reset] Suppression des donnees existantes...")
        db.query(GeoAvailability).delete()
        db.query(GeoSupplier).delete()
        db.query(GeoDock).delete()
        db.commit()

        # Parser dans l'ordre
        n_sup, skip_sup = parse_suppliers(sheet_suppliers, db)
        n_dock, skip_dock = parse_docks(sheet_docks, db)
        n_avail, _ = parse_availability(sheet_availability, db)

        print("\n" + "=" * 60)
        print("RESUME")
        print("=" * 60)
        print(f"Suppliers inseres   : {n_sup} (skip: {skip_sup})")
        print(f"Docks inseres       : {n_dock} (skip: {skip_dock})")
        print(f"Availability inseres: {n_avail}")
        print("=" * 60)
        print("OK")

    except Exception as e:
        db.rollback()
        print(f"\n[FATAL] Erreur durant le seed : {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        db.close()


if __name__ == "__main__":
    main()