"""
services/geo_stock_service.py
-----------------------------
Parser le fichier de stock client (Tableau Croise Dynamique xlsx).

Format attendu :
- Lignes 1-3 : entetes de filtres (Booking Mode, Week, Gestion) - ignorees
- Ligne 5 : header "Valeurs" - ignoree
- Ligne 6 : VRAIS headers (Container, CU, Packaging Code, ..., STOCK DISPO, ...)
- Lignes 7+ : data

Politique :
- Ignorer Container = ' ' (espace) ou vide ou 'Grand Total'
- Ignorer si Packaging Code vide
- Ignorer si Stock dispo NULL
"""

from io import BytesIO
import openpyxl


# Headers attendus (case-insensitive, espaces-insensitive)
HEADER_CONTAINER = "container service center"
HEADER_PACKAGING = "packaging code"
HEADER_CU = "cu"
HEADER_STOCK = "stock dispo"


def _normalize(s):
    """Normalize string : lowercase, strip, collapse spaces, remove newlines."""
    if s is None:
        return ""
    return " ".join(str(s).lower().replace("\n", " ").split())


def _to_float(v):
    """Convertit en float, retourne None si invalide."""
    if v is None or v == "":
        return None
    try:
        if isinstance(v, str):
            v = v.replace(",", ".").strip()
        f = float(v)
        if f != f:  # NaN check
            return None
        return f
    except (ValueError, TypeError):
        return None


def parse_stock_xlsx(file_bytes: bytes) -> tuple[list[dict], dict]:
    """
    Parse le fichier xlsx et retourne (entries, stats).
    
    entries: list of {dock_name, packaging_code, qty_available, cu}
    stats:   dict avec compteurs (total_rows, inserted, skipped_*)
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active  # On prend la 1ere feuille (le format est mono-feuille)

    # === Trouver la ligne d'entete (cherche la ligne contenant "Container Service Center") ===
    header_row_idx = None
    headers_map = {}  # {normalized_header: col_idx (0-indexed)}
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        # Cherche dans cette ligne si on a "container service center"
        norm_row = [_normalize(c) for c in row]
        if HEADER_CONTAINER in norm_row:
            header_row_idx = row_idx
            for col_idx, val in enumerate(norm_row):
                if val:
                    headers_map[val] = col_idx
            break

    if header_row_idx is None:
        raise ValueError("Header 'Container Service Center' introuvable dans les 20 premieres lignes")

    # Verifier les colonnes requises
    if HEADER_PACKAGING not in headers_map:
        raise ValueError(f"Colonne '{HEADER_PACKAGING}' introuvable")
    if HEADER_STOCK not in headers_map:
        raise ValueError(f"Colonne '{HEADER_STOCK}' introuvable")

    col_container = headers_map[HEADER_CONTAINER]
    col_packaging = headers_map[HEADER_PACKAGING]
    col_stock = headers_map[HEADER_STOCK]
    col_cu = headers_map.get(HEADER_CU)  # optionnel

    # === Parse les lignes de data ===
    entries = []
    stats = {
        "total_rows": 0,
        "inserted": 0,
        "skipped_no_container": 0,
        "skipped_grand_total": 0,
        "skipped_no_packaging": 0,
        "skipped_no_stock": 0,
        "header_row_idx": header_row_idx,
    }

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        # Ignorer lignes totalement vides
        if all(c is None for c in row):
            continue

        stats["total_rows"] += 1

        # Container : doit etre rempli (pas vide ni juste un espace)
        container_raw = row[col_container]
        container = str(container_raw).strip() if container_raw else ""
        if not container:
            stats["skipped_no_container"] += 1
            continue
        if container.lower() == "grand total":
            stats["skipped_grand_total"] += 1
            continue

        # Packaging : doit etre rempli
        packaging_raw = row[col_packaging]
        if packaging_raw is None or str(packaging_raw).strip() == "":
            stats["skipped_no_packaging"] += 1
            continue
        packaging = str(packaging_raw).strip()

        # Stock dispo : doit etre numerique (peut etre negatif)
        stock = _to_float(row[col_stock])
        if stock is None:
            stats["skipped_no_stock"] += 1
            continue

        # CU optionnel
        cu = None
        if col_cu is not None:
            cu_raw = row[col_cu]
            if cu_raw:
                cu = str(cu_raw).strip()

        entries.append({
            "dock_name": container,
            "packaging_code": packaging,
            "qty_available": stock,
            "cu": cu,
        })
        stats["inserted"] += 1

    wb.close()
    return entries, stats